# -*- encoding: utf-8 -*-
"""
@File    : file_operate.py
@Time    : 2021/10/11 16:13
@Author  : Coco
"""
# -*- encoding: utf-8 -*-
"""
@File    : file_operate.py
@Time    : 2021/9/29 14:48
@Author  : Coco
"""
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment


# 创建表格
def create_xlsx(file_path, para_list):
    """
    用于创建固定格式的Excel表格
    :param para_list:
    :param file_path: 创建表格的路径及名称
    :return:
    """
    workbook = Workbook()
    # 激活 worksheet
    worksheet = workbook.active
    sheet = workbook.worksheets[0]
    # 设置头部信息
    worksheet.column_dimensions['A'].width = 20.0
    worksheet.column_dimensions['B'].width = 20.0
    worksheet.column_dimensions['C'].width = 20.0
    worksheet.column_dimensions['D'].width = 20.0
    worksheet.column_dimensions['E'].width = 20.0
    worksheet.column_dimensions['F'].width = 20.0
    worksheet.column_dimensions['G'].width = 20.0
    sheet['B1'].alignment = Alignment(horizontal='left', vertical='center')
    worksheet['A1'] = '报告生成时间：'
    sheet.merge_cells('B1:C1')
    worksheet['B1'] = datetime.now()
    worksheet['A2'] = "ID"
    for i in range(len(para_list)):
        print(para_list[i])
        worksheet.cell(2, i + 2).value = para_list[i]
    # 保存
    workbook.save(file_path)


# 写入数据
def write_data(file_path, data_list, col, *args):
    if args:
        # if not os.path.exists(file_path):
        para_list = args[0]
        print(para_list)
        create_xlsx(file_path, para_list)
    wb = load_workbook(file_path)
    ws = wb.active
    ws['B1'] = datetime.now()
    for row in range(3, len(data_list) + 3):
        ws.cell(row, col).value = data_list[row - 3]
    wb.save(file_path)


def read_data_list(file_path, col):
    data_list = []
    wb = load_workbook(file_path)
    ws = wb.active
    for row in range(3, ws.max_row + 3):
        data_list.append(ws.cell(row, col).value)
    return data_list


if __name__ == '__main__':
    create_xlsx('../data/test2.xlsx', [])
    # write_data('../data/test2.xlsx', [i for i in range(100)], 1, "123")
